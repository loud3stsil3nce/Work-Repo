import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Database and table info
DB_PATH = "informationUpload.db"  # Update with actual database path
TABLE_NAME = "alerts"  # Update with actual table name
OUTPUT_EXCEL = "outage_report.xlsx"

# Connect to SQLite and fetch data
conn = sqlite3.connect(DB_PATH)
query = f"SELECT alert_time, duration, monitoring_server FROM {TABLE_NAME}"
df = pd.read_sql(query, conn)
conn.close()

# Convert alert_time to date-only format
df["alert_time"] = pd.to_datetime(df["alert_time"]).dt.date

# Aggregate outage durations per server per day
grouped_df = df.groupby(["alert_time", "monitoring_server"])["duration"].sum().unstack(fill_value=0)

# Write to Excel
grouped_df.to_excel(OUTPUT_EXCEL, sheet_name="Outages")

# Load workbook and select sheet
wb = load_workbook(OUTPUT_EXCEL)
ws = wb["Outages"]

# Define data range for the chart
min_col, max_col = 2, ws.max_column
min_row, max_row = 1, ws.max_row

# Create a bar chart
chart = BarChart()
chart.title = "Total Outage Duration by Server"
chart.x_axis.title = "Date"
chart.y_axis.title = "Duration (ms)"

# Reference data range
data = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row, max_col=max_col)
categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.type = "col"  # Column chart (stacked)

# Insert the chart in Excel
ws.add_chart(chart, "E5")

# Save the updated Excel file
wb.save(OUTPUT_EXCEL)

print(f"✅ Excel file with a graph has been created: {OUTPUT_EXCEL}")